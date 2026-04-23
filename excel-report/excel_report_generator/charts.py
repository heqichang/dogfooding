from openpyxl.chart import (
    BarChart, LineChart, Reference,
    BarChart3D, LineChart3D
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Color
from typing import List, Dict, Any, Optional, Tuple
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ChartGenerator:
    DEFAULT_COLORS = [
        "4472C4", "ED7D31", "70AD47", "FFC000",
        "5B9BD5", "7030A0", "C65911", "00B0F0"
    ]

    @staticmethod
    def create_bar_chart(
        sheet,
        data_range: str,
        categories_range: str,
        title: str = "",
        x_axis_title: str = "",
        y_axis_title: str = "",
        chart_type: str = "col",
        style: int = 10,
        show_legend: bool = True,
        show_data_labels: bool = False,
        colors: Optional[List[str]] = None,
        three_dimensional: bool = False
    ):
        if three_dimensional:
            chart = BarChart3D()
        else:
            chart = BarChart()

        chart.type = chart_type
        chart.style = style
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title

        data = Reference(sheet, range_string=data_range)
        categories = Reference(sheet, range_string=categories_range)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        if not show_legend:
            chart.legend = None

        if show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

        if colors:
            ChartGenerator._apply_series_colors(chart, colors)

        logger.info(f"柱状图已创建: {title if title else '未命名'}")
        return chart

    @staticmethod
    def create_line_chart(
        sheet,
        data_range: str,
        categories_range: str,
        title: str = "",
        x_axis_title: str = "",
        y_axis_title: str = "",
        style: int = 10,
        show_legend: bool = True,
        show_data_labels: bool = False,
        show_markers: bool = False,
        smooth_lines: bool = False,
        colors: Optional[List[str]] = None,
        three_dimensional: bool = False
    ):
        if three_dimensional:
            chart = LineChart3D()
        else:
            chart = LineChart()

        chart.style = style
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title

        data = Reference(sheet, range_string=data_range)
        categories = Reference(sheet, range_string=categories_range)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        if not show_legend:
            chart.legend = None

        if show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

        for series in chart.series:
            if show_markers:
                series.marker.symbol = "circle"
                series.marker.size = 5
            if smooth_lines:
                series.smooth = True

        if colors:
            ChartGenerator._apply_series_colors(chart, colors)

        logger.info(f"折线图已创建: {title if title else '未命名'}")
        return chart

    @staticmethod
    def _apply_series_colors(chart, colors: List[str]):
        for idx, series in enumerate(chart.series):
            color = colors[idx % len(colors)]
            if hasattr(series, 'graphicalProperties'):
                series.graphicalProperties.solidFill = color
            elif hasattr(series, 'spPr'):
                series.spPr.solidFill = color

    @staticmethod
    def add_chart_to_sheet(
        sheet,
        chart,
        position: str = "A1",
        width: int = 15,
        height: int = 10
    ):
        chart.width = width
        chart.height = height
        sheet.add_chart(chart, position)
        logger.info(f"图表已添加到位置: {position}")

    @staticmethod
    def create_chart_from_data(
        sheet,
        data: List[Dict[str, Any]],
        chart_type: str = "bar",
        title: str = "",
        x_axis_title: str = "",
        y_axis_title: str = "",
        category_col: str = None,
        value_cols: List[str] = None,
        position: str = "A1",
        style: int = 10,
        show_legend: bool = True,
        show_data_labels: bool = False,
        colors: Optional[List[str]] = None
    ) -> Tuple[Any, str]:
        if not data:
            raise ValueError("数据不能为空")

        headers = list(data[0].keys())
        if category_col is None:
            category_col = headers[0]
        if value_cols is None:
            value_cols = [h for h in headers if h != category_col]

        start_row = sheet.max_row + 2
        start_col = 1

        for col_idx, header in enumerate([category_col] + value_cols, start=start_col):
            sheet.cell(row=start_row, column=col_idx, value=header)

        for row_idx, row_data in enumerate(data, start=start_row + 1):
            sheet.cell(row=row_idx, column=start_col, value=row_data[category_col])
            for col_idx, value_col in enumerate(value_cols, start=start_col + 1):
                sheet.cell(row=row_idx, column=col_idx, value=row_data[value_col])

        end_row = start_row + len(data)
        end_col = start_col + len(value_cols)

        from openpyxl.utils import get_column_letter
        data_range = f"{get_column_letter(start_col + 1)}{start_row}:{get_column_letter(end_col)}{end_row}"
        categories_range = f"{get_column_letter(start_col)}{start_row + 1}:{get_column_letter(start_col)}{end_row}"

        if chart_type.lower() == "bar":
            chart = ChartGenerator.create_bar_chart(
                sheet=sheet,
                data_range=data_range,
                categories_range=categories_range,
                title=title,
                x_axis_title=x_axis_title,
                y_axis_title=y_axis_title,
                style=style,
                show_legend=show_legend,
                show_data_labels=show_data_labels,
                colors=colors
            )
        elif chart_type.lower() == "line":
            chart = ChartGenerator.create_line_chart(
                sheet=sheet,
                data_range=data_range,
                categories_range=categories_range,
                title=title,
                x_axis_title=x_axis_title,
                y_axis_title=y_axis_title,
                style=style,
                show_legend=show_legend,
                show_data_labels=show_data_labels,
                colors=colors
            )
        else:
            raise ValueError(f"不支持的图表类型: {chart_type}")

        ChartGenerator.add_chart_to_sheet(sheet, chart, position=position)

        return chart, f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

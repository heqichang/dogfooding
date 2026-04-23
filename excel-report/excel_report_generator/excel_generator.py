from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill,
    NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Union
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class StyleConfig:
    DEFAULT_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
    DEFAULT_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    DEFAULT_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DEFAULT_CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    DEFAULT_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def create_header_style(
        font: Font = None,
        fill: PatternFill = None,
        alignment: Alignment = None,
        border: Border = None
    ) -> NamedStyle:
        style = NamedStyle(name="header_style")
        style.font = font or StyleConfig.DEFAULT_HEADER_FONT
        style.fill = fill or StyleConfig.DEFAULT_HEADER_FILL
        style.alignment = alignment or StyleConfig.DEFAULT_HEADER_ALIGNMENT
        style.border = border or StyleConfig.DEFAULT_BORDER
        return style

    @staticmethod
    def create_data_style(
        font: Font = None,
        alignment: Alignment = None,
        border: Border = None,
        number_format: str = None
    ) -> NamedStyle:
        style = NamedStyle(name="data_style")
        style.font = font or Font(size=11)
        style.alignment = alignment or StyleConfig.DEFAULT_CELL_ALIGNMENT
        style.border = border or StyleConfig.DEFAULT_BORDER
        if number_format:
            style.number_format = number_format
        return style


class ExcelGenerator:
    def __init__(self):
        self.workbook = Workbook()
        self._remove_default_sheet()
        self.header_style = StyleConfig.create_header_style()
        self.data_style = StyleConfig.create_data_style()
        self._register_styles()
        logger.info("Excel生成器已初始化")

    def _remove_default_sheet(self):
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

    def _register_styles(self):
        try:
            self.workbook.add_named_style(self.header_style)
        except ValueError:
            pass
        try:
            self.workbook.add_named_style(self.data_style)
        except ValueError:
            pass

    def create_sheet(
        self,
        sheet_name: str,
        data: List[Dict[str, Any]],
        include_headers: bool = True,
        start_row: int = 1,
        start_col: int = 1,
        header_style: Optional[NamedStyle] = None,
        data_style: Optional[NamedStyle] = None
    ) -> Any:
        if not data:
            logger.warning(f"数据为空，跳过创建Sheet: {sheet_name}")
            return None

        if sheet_name in self.workbook.sheetnames:
            sheet_name = f"{sheet_name}_{len(self.workbook.sheetnames)}"

        sheet = self.workbook.create_sheet(title=sheet_name)
        h_style = header_style or self.header_style
        d_style = data_style or self.data_style

        if not data:
            return sheet

        headers = list(data[0].keys()) if include_headers else []

        if include_headers:
            for col_idx, header in enumerate(headers, start=start_col):
                cell = sheet.cell(row=start_row, column=col_idx, value=header)
                cell.style = h_style
            data_start_row = start_row + 1
        else:
            data_start_row = start_row

        for row_idx, row_data in enumerate(data, start=data_start_row):
            for col_idx, header in enumerate(headers or row_data.keys(), start=start_col):
                value = row_data.get(header) if headers else list(row_data.values())[col_idx - start_col]
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.style = d_style

        self._auto_fit_columns(sheet, start_col, start_row, len(headers or data[0]))
        logger.info(f"Sheet '{sheet_name}' 已创建，包含 {len(data)} 行数据")
        return sheet

    def _auto_fit_columns(self, sheet, start_col: int, start_row: int, num_cols: int):
        for col_idx in range(start_col, start_col + num_cols):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in sheet[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def merge_cells(
        self,
        sheet: Any,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        value: Any = None,
        style: Optional[NamedStyle] = None
    ):
        sheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col
        )
        cell = sheet.cell(row=start_row, column=start_col)
        if value is not None:
            cell.value = value
        if style:
            cell.style = style
        logger.info(f"合并单元格: ({start_row},{start_col}) 到 ({end_row},{end_col})")

    def apply_style_to_range(
        self,
        sheet: Any,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        style: NamedStyle
    ):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).style = style

    def set_cell_value(
        self,
        sheet: Any,
        row: int,
        col: int,
        value: Any,
        style: Optional[NamedStyle] = None
    ):
        cell = sheet.cell(row=row, column=col, value=value)
        if style:
            cell.style = style

    def get_sheet(self, sheet_name: str):
        return self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else None

    def save(self, filename: str):
        self.workbook.save(filename)
        logger.info(f"Excel文件已保存: {filename}")

    def save_to_bytes(self):
        import io
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

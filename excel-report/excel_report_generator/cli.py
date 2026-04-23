import argparse
import sys
from pathlib import Path
from typing import Optional, Dict, Any

from excel_report_generator.database import DatabaseConnector
from excel_report_generator.excel_generator import ExcelGenerator, StyleConfig
from excel_report_generator.charts import ChartGenerator
from excel_report_generator.templates import TemplateRenderer
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ReportGeneratorCLI:
    def __init__(self):
        self.parser = self._create_parser()

    def _create_parser(self) -> argparse.ArgumentParser:
        parser = argparse.ArgumentParser(
            description='Excel报表生成工具 - 支持多Sheet、图表、样式和模板',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog='''
示例用法:
  # 使用模板生成报表
  excel-report generate --template config.yaml --db sqlite:///data.db --output report.xlsx

  # 直接查询数据库生成报表
  excel-report query --db sqlite:///data.db --queries "SELECT * FROM sales" --sheets "销售数据"

  # 生成示例模板
  excel-report init --output template.yaml

  # 测试数据库连接
  excel-report test --db sqlite:///data.db
            '''
        )

        subparsers = parser.add_subparsers(dest='command', help='可用命令')

        generate_parser = subparsers.add_parser('generate', help='使用模板生成报表')
        generate_parser.add_argument('--template', '-t', required=True, help='模板文件路径 (YAML/JSON)')
        generate_parser.add_argument('--db', '-d', help='数据库连接字符串')
        generate_parser.add_argument('--output', '-o', help='输出Excel文件路径')
        generate_parser.add_argument('--context', '-c', nargs='*', help='上下文变量 (key=value)')

        query_parser = subparsers.add_parser('query', help='直接查询数据库生成报表')
        query_parser.add_argument('--db', '-d', required=True, help='数据库连接字符串')
        query_parser.add_argument('--queries', '-q', nargs='+', required=True, help='SQL查询语句列表')
        query_parser.add_argument('--sheets', '-s', nargs='+', help='Sheet名称列表 (与查询对应)')
        query_parser.add_argument('--output', '-o', default='report.xlsx', help='输出Excel文件路径')
        query_parser.add_argument('--charts', action='store_true', help='为数据自动生成图表')

        init_parser = subparsers.add_parser('init', help='生成示例模板配置文件')
        init_parser.add_argument('--output', '-o', default='template.yaml', help='输出模板文件路径')
        init_parser.add_argument('--format', '-f', choices=['yaml', 'json'], default='yaml', help='文件格式')

        test_parser = subparsers.add_parser('test', help='测试数据库连接')
        test_parser.add_argument('--db', '-d', required=True, help='数据库连接字符串')

        return parser

    def _parse_context(self, context_args) -> Dict[str, Any]:
        context = {}
        if context_args:
            for arg in context_args:
                if '=' in arg:
                    key, value = arg.split('=', 1)
                    context[key.strip()] = value.strip()
        return context

    def run(self, args=None):
        parsed_args = self.parser.parse_args(args)

        if parsed_args.command == 'generate':
            return self._handle_generate(parsed_args)
        elif parsed_args.command == 'query':
            return self._handle_query(parsed_args)
        elif parsed_args.command == 'init':
            return self._handle_init(parsed_args)
        elif parsed_args.command == 'test':
            return self._handle_test(parsed_args)
        else:
            self.parser.print_help()
            return 1

    def _handle_generate(self, args) -> int:
        try:
            renderer = TemplateRenderer()
            template = renderer.load_template(args.template)

            context = self._parse_context(args.context)
            renderer.set_context(context)

            errors = renderer.validate_template()
            if errors:
                for error in errors:
                    logger.error(f"模板验证错误: {error}")
                return 1

            excel_gen = ExcelGenerator()
            db_connector = None

            if args.db:
                db_connector = DatabaseConnector(args.db)
                if not db_connector.test_connection():
                    logger.error("数据库连接失败")
                    return 1

            sheet_configs = renderer.get_sheet_configs()

            for sheet_config in sheet_configs:
                sheet_name = sheet_config.get('name', 'Sheet')
                sheet_name = renderer.render_value(sheet_name)

                data = None
                if 'query' in sheet_config and db_connector:
                    query = renderer.render_value(sheet_config['query'])
                    logger.info(f"执行查询: {query[:100]}...")
                    data = db_connector.execute_query(query)
                elif 'data' in sheet_config:
                    data = renderer.render_value(sheet_config['data'])
                elif 'static_data' in sheet_config:
                    data = sheet_config['static_data']

                if data:
                    sheet = excel_gen.create_sheet(
                        sheet_name=sheet_name,
                        data=data,
                        include_headers=sheet_config.get('include_headers', True)
                    )

                    if 'merge_cells' in sheet_config and sheet:
                        for merge_config in sheet_config['merge_cells']:
                            excel_gen.merge_cells(
                                sheet=sheet,
                                start_row=merge_config.get('start_row', 1),
                                end_row=merge_config.get('end_row', 1),
                                start_col=merge_config.get('start_col', 1),
                                end_col=merge_config.get('end_col', 1),
                                value=merge_config.get('value')
                            )

                    if 'charts' in sheet_config and sheet and len(data) > 1:
                        for chart_config in sheet_config['charts']:
                            chart_type = chart_config.get('type', 'bar')
                            category_col = chart_config.get('x_axis')
                            value_cols = chart_config.get('y_axes', [])

                            try:
                                ChartGenerator.create_chart_from_data(
                                    sheet=sheet,
                                    data=data,
                                    chart_type=chart_type,
                                    title=chart_config.get('title', ''),
                                    category_col=category_col,
                                    value_cols=value_cols if value_cols else None,
                                    position=chart_config.get('position', 'A10'),
                                    show_data_labels=chart_config.get('show_data_labels', False)
                                )
                            except Exception as e:
                                logger.warning(f"图表创建失败: {e}")

            output_path = args.output or renderer.generate_output_name()
            excel_gen.save(output_path)
            logger.info(f"报表已生成: {output_path}")

            if db_connector:
                db_connector.disconnect()

            return 0

        except Exception as e:
            logger.error(f"生成报表失败: {str(e)}", exc_info=True)
            return 1

    def _handle_query(self, args) -> int:
        try:
            db_connector = DatabaseConnector(args.db)
            if not db_connector.test_connection():
                logger.error("数据库连接失败")
                return 1

            excel_gen = ExcelGenerator()

            sheet_names = args.sheets or [f"Sheet{i+1}" for i in range(len(args.queries))]

            if len(sheet_names) < len(args.queries):
                sheet_names.extend([f"Sheet{i+1}" for i in range(len(sheet_names), len(args.queries))])

            for idx, (query, sheet_name) in enumerate(zip(args.queries, sheet_names)):
                logger.info(f"执行查询 {idx+1}: {query[:100]}...")
                data = db_connector.execute_query(query)

                if data:
                    sheet = excel_gen.create_sheet(
                        sheet_name=sheet_name,
                        data=data
                    )

                    if args.charts and sheet and len(data) > 1:
                        try:
                            ChartGenerator.create_chart_from_data(
                                sheet=sheet,
                                data=data,
                                chart_type='bar',
                                title=f"{sheet_name} 统计",
                                position='A10'
                            )
                        except Exception as e:
                            logger.warning(f"图表创建失败: {e}")
                else:
                    logger.warning(f"查询 {idx+1} 无数据返回")

            excel_gen.save(args.output)
            logger.info(f"报表已生成: {args.output}")

            db_connector.disconnect()
            return 0

        except Exception as e:
            logger.error(f"查询生成报表失败: {str(e)}", exc_info=True)
            return 1

    def _handle_init(self, args) -> int:
        try:
            TemplateRenderer.save_sample_template(args.output, args.format)
            logger.info(f"示例模板已生成: {args.output}")
            return 0
        except Exception as e:
            logger.error(f"生成模板失败: {str(e)}")
            return 1

    def _handle_test(self, args) -> int:
        try:
            db_connector = DatabaseConnector(args.db)
            if db_connector.test_connection():
                logger.info("数据库连接成功!")
                return 0
            else:
                logger.error("数据库连接失败")
                return 1
        except Exception as e:
            logger.error(f"连接测试失败: {str(e)}")
            return 1


def main():
    cli = ReportGeneratorCLI()
    sys.exit(cli.run())


if __name__ == '__main__':
    main()

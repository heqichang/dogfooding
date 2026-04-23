from excel_report_generator.database import DatabaseConnector
from excel_report_generator.excel_generator import ExcelGenerator, StyleConfig
from excel_report_generator.charts import ChartGenerator
from excel_report_generator.templates import TemplateRenderer
from openpyxl.styles import Font, PatternFill, Alignment
import os


def demo_database_connection():
    print("=" * 60)
    print("演示1: 数据库连接和查询")
    print("=" * 60)

    db_path = os.path.join(os.path.dirname(__file__), "sample.db")
    if not os.path.exists(db_path):
        print(f"警告: 数据库文件不存在: {db_path}")
        print("请先运行 create_sample_db.py 创建示例数据库")
        return None

    with DatabaseConnector(f"sqlite:///{db_path}") as db:
        print("\n1. 测试数据库连接...")
        if db.test_connection():
            print("✓ 数据库连接成功")
        else:
            print("✗ 数据库连接失败")
            return None

        print("\n2. 执行查询获取月度销售数据...")
        sales_data = db.execute_query("SELECT * FROM monthly_sales LIMIT 5")
        for row in sales_data:
            print(f"  {row}")

        print("\n3. 执行联合查询获取员工信息...")
        employee_data = db.execute_query("""
            SELECT e.name, d.name as dept, e.position, e.salary
            FROM employees e
            JOIN departments d ON e.department_id = d.id
            LIMIT 5
        """)
        for row in employee_data:
            print(f"  {row}")

        print("\n✓ 数据库操作演示完成")
        return db_path


def demo_excel_generation():
    print("\n" + "=" * 60)
    print("演示2: Excel生成 - 多Sheet、样式、合并单元格")
    print("=" * 60)

    excel = ExcelGenerator()

    sample_data = [
        {"月份": "2024-01", "销售额": 125000, "利润": 31250, "订单数": 125},
        {"月份": "2024-02", "销售额": 98000, "利润": 24500, "订单数": 98},
        {"月份": "2024-03", "销售额": 156000, "利润": 39000, "订单数": 156},
        {"月份": "2024-04", "销售额": 142000, "利润": 35500, "订单数": 142},
        {"月份": "2024-05", "销售额": 168000, "利润": 42000, "订单数": 168},
    ]

    print("\n1. 创建第一个Sheet: 销售数据")
    sheet1 = excel.create_sheet("销售数据", sample_data)
    print(f"✓ Sheet '销售数据' 已创建，包含 {len(sample_data)} 行数据")

    print("\n2. 创建自定义样式...")
    title_font = Font(bold=True, size=16, color="4472C4")
    title_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    print("✓ 自定义样式已创建")

    print("\n3. 合并单元格并添加标题...")
    excel.merge_cells(
        sheet=sheet1,
        start_row=1,
        end_row=1,
        start_col=1,
        end_col=4,
        value="2024年上半年销售统计"
    )
    print("✓ 单元格已合并，标题已添加")

    product_data = [
        {"产品名称": "笔记本电脑", "类别": "电子产品", "价格": 5999.99, "库存": 100},
        {"产品名称": "无线鼠标", "类别": "配件", "价格": 89.99, "库存": 500},
        {"产品名称": "机械键盘", "类别": "配件", "价格": 399.99, "库存": 200},
        {"产品名称": "显示器", "类别": "电子产品", "价格": 1999.99, "库存": 150},
    ]

    print("\n4. 创建第二个Sheet: 产品列表")
    sheet2 = excel.create_sheet("产品列表", product_data)
    print(f"✓ Sheet '产品列表' 已创建，包含 {len(product_data)} 行数据")

    output_path = os.path.join(os.path.dirname(__file__), "demo_output.xlsx")
    excel.save(output_path)
    print(f"\n✓ Excel文件已保存: {output_path}")

    return output_path, sample_data, sheet1


def demo_chart_generation(sample_data, sheet):
    print("\n" + "=" * 60)
    print("演示3: 图表生成 - 柱状图和折线图")
    print("=" * 60)

    if sheet is None or sample_data is None:
        print("警告: 无法进行图表演示，缺少必要的数据")
        return

    print("\n1. 从数据自动生成柱状图...")
    chart, data_range = ChartGenerator.create_chart_from_data(
        sheet=sheet,
        data=sample_data,
        chart_type="bar",
        title="月度销售额对比",
        x_axis_title="月份",
        y_axis_title="金额",
        category_col="月份",
        value_cols=["销售额", "利润"],
        position="A10",
        show_data_labels=True
    )
    print(f"✓ 柱状图已创建，数据范围: {data_range}")

    print("\n2. 创建折线图...")
    line_chart, line_range = ChartGenerator.create_chart_from_data(
        sheet=sheet,
        data=sample_data,
        chart_type="line",
        title="订单数趋势",
        x_axis_title="月份",
        y_axis_title="订单数",
        category_col="月份",
        value_cols=["订单数"],
        position="A30",
        show_markers=True,
        smooth_lines=True
    )
    print(f"✓ 折线图已创建，数据范围: {line_range}")

    output_path = os.path.join(os.path.dirname(__file__), "demo_charts.xlsx")
    sheet.parent.save(output_path)
    print(f"\n✓ 包含图表的Excel文件已保存: {output_path}")


def demo_template_rendering():
    print("\n" + "=" * 60)
    print("演示4: 模板渲染")
    print("=" * 60)

    renderer = TemplateRenderer()

    print("\n1. 加载模板文件...")
    template_path = os.path.join(os.path.dirname(__file__), "sample_template.yaml")
    if os.path.exists(template_path):
        template = renderer.load_template(template_path)
        print(f"✓ 模板已加载: {template_path}")
        print(f"  - Sheet数量: {len(template.get('sheets', []))}")
    else:
        print(f"警告: 模板文件不存在: {template_path}")

    print("\n2. 测试变量渲染...")
    test_context = {
        "company": "示例科技有限公司",
        "year": 2024,
        "department": {"name": "销售部", "manager": "张三"}
    }
    renderer.set_context(test_context)

    test_value = "{{company}} - {{year}}年度报告 - 部门: {{department.name}}"
    rendered = renderer.render_value(test_value)
    print(f"  原始值: {test_value}")
    print(f"  渲染后: {rendered}")

    print("\n3. 验证模板...")
    errors = renderer.validate_template()
    if errors:
        for error in errors:
            print(f"  ✗ {error}")
    else:
        print("  ✓ 模板验证通过")

    print("\n4. 生成输出文件名...")
    output_name = renderer.generate_output_name("test_report")
    print(f"  生成的文件名: {output_name}")


def run_all_demos():
    print("\n" + "=" * 60)
    print("Excel报表生成器 - 完整演示")
    print("=" * 60)

    db_path = demo_database_connection()

    output_path, sample_data, sheet = demo_excel_generation()

    demo_chart_generation(sample_data, sheet)

    demo_template_rendering()

    print("\n" + "=" * 60)
    print("所有演示完成！")
    print("=" * 60)
    print("\n生成的文件:")
    if output_path and os.path.exists(output_path):
        print(f"  - {output_path}")
    charts_path = os.path.join(os.path.dirname(__file__), "demo_charts.xlsx")
    if os.path.exists(charts_path):
        print(f"  - {charts_path}")
    print("\n您可以用Excel打开这些文件查看效果。")


if __name__ == "__main__":
    run_all_demos()
